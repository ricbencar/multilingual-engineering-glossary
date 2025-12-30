"""
====================================================================================================
  SCRIPT NAME: High-Performance Multi-Language Excel Translator (Stage 1)
====================================================================================================

  --------------------------------------------------------------------------------------------------
  1. OVERVIEW
  --------------------------------------------------------------------------------------------------
  This script is the "Stage 1" processor for multilingual projects. It automates the translation 
  of a source Excel file ('english.xlsx') into up to 30 global languages, producing a highly 
  formatted 'glossary.xlsx' file.

  It solves the "Excel Tofu Problem" (empty boxes like ☐☐☐) by intelligently detecting languages 
  and embedding the specific Google Noto font families required to render them correctly.

  --------------------------------------------------------------------------------------------------
  2. INSTALLATION (PYTHON LIBRARIES)
  --------------------------------------------------------------------------------------------------
  Open your terminal/command prompt and run:
      pip install pandas openpyxl deep-translator

  --------------------------------------------------------------------------------------------------
  3. INPUT FILE REQUIREMENTS
  --------------------------------------------------------------------------------------------------
  Create a file named 'english.xlsx' in the same folder as this script.
  It MUST contain at least these two columns:
    - English_word  (The term to translate)
    - English_descr (The description to translate)
   
  (Optional: Add a 'Category' column to group terms).

  --------------------------------------------------------------------------------------------------
  4. FONT SETUP (CRITICAL)
  --------------------------------------------------------------------------------------------------
  For Excel and the PDF Generator to work, you MUST have these exact fonts.
  Missing fonts = Empty boxes (☐) or script crashes.

  [FOLDER STRUCTURE]
  Ensure your 'fonts' folder contains ALL of these files to support the 30-language list:

      /your_project/
        ├── get_fonts.py             # Stage 0: Fonts Downloader
        ├── script_translate.py      # Stage 1: Translator
        ├── script_glossary.py       # Stage 2: PDF Renderer (Glossary)
        ├── script_table.py          # Stage 2: PDF Renderer (Table)
        ├── english.xlsx             # Source Input
        ├── glossary.xlsx            # Intermediate Output
        └── fonts/
              │  # --- 1. CORE FONTS (Latin/Cyrillic & Headers) ---
              ├── NotoSansLiving-Regular.ttf      <-- Base text (English, French, Spanish, Turkish, etc.)
              ├── NotoSans-Bold.ttf               <-- REQUIRED for Headers/Titles in PDF generation
              │
              │  # --- 2. EAST ASIAN (CJK) ---
              ├── NotoSansCJK.ttc                 <-- "Super Font" for Chinese, Japanese, & Korean
              │                                       (Note: .ttc is a Collection file. Do not use .otf)
              │
              │  # --- 3. MIDDLE EASTERN (RTL) ---
              ├── NotoSansArabic-Regular.ttf      <-- Arabic, Persian, & Urdu (Fallback)
              ├── NotoNastaliqUrdu-Regular.ttf    <-- Urdu (Preferred "Cascading" style)
              │
              │  # --- 4. SOUTH ASIAN (Indic Scripts) ---
              ├── NotoSansDevanagari-Regular.ttf  <-- Hindi & Marathi
              ├── NotoSansBengali-Regular.ttf     <-- Bengali
              ├── NotoSansGujarati-Regular.ttf    <-- Gujarati
              ├── NotoSansTamil-Regular.ttf       <-- Tamil
              ├── NotoSansTelugu-Regular.ttf      <-- Telugu
              ├── NotoSansGurmukhi-Regular.ttf    <-- Western Punjabi (Google outputs Gurmukhi script)
              │
              │  # --- 5. SOUTHEAST ASIAN ---
              ├── NotoSansThai-Regular.ttf        <-- Thai
              ├── NotoSansJavanese-Regular.ttf    <-- Javanese

  [REQUIRED FONTS TO DOWNLOAD]
  To guarantee zero "Tofu" (☐☐☐) and perfect rendering in both Excel and PDF, you must 
  download the specific font files listed below.

  A. THE "CORE" FONTS (Latin, Cyrillic, Greek & Headers)
     * Filenames: "NotoSansLiving-Regular.ttf" AND "NotoSans-Bold.ttf"
     * Source:    https://github.com/notofonts/notofonts.github.io/tree/main/megamerge
     * Why:       
       1. "Regular" (Living): This specific "Mega-Merge" version covers ~80% of languages (English, 
          French, Russian, etc.) in a single file. It fixes rendering issues for Turkish (İ, ş) 
          and Vietnamese (stacked diacritics) that standard Arial often breaks.
       2. "Bold": MANDATORY for the PDF generator. Without 'NotoSans-Bold.ttf', section headers 
          (e.g., "CATEGORY") will fail to render, causing the script to crash or print blank headers.

  B. THE "CJK" SUPER-FONT (Chinese, Japanese, Korean)
     * Filename:  "NotoSerifCJK.ttc" (or NotoSansCJK.ttc)
     * Source:    https://github.com/notofonts/noto-cjk
     * Why:       
       1. Scale: CJK languages require >65,000 glyphs. Standard fonts do not contain them all.
       2. Efficiency: The ".ttc" (TrueType Collection) format bundles Simplified Chinese (SC), 
          Traditional Chinese (TC), Japanese (JP), and Korean (KR) into one efficient file.
       3. Compatibility: This script is tuned to detect the "TTC" collection. Using individual 
          ".otf" files may result in Excel failing to recognize the font family.

  C. MIDDLE EASTERN (Right-to-Left Scripts)
     * Filenames: 
       1. "NotoSansArabic-Regular.ttf" (Essential for Arabic, Persian, & Standard Urdu)
       2. "NotoNastaliqUrdu-Regular.ttf" (Recommended for Urdu aesthetics)
     * Source:    https://www.google.com/get/noto/
     * Why:       
       1. Shaping: Arabic letters change shape based on position (Start/Middle/End). Standard 
          fonts often break these "ligatures," leaving letters disconnected (e.g., م ك instead of مك).
       2. Style: Urdu users prefer "Nastaliq" (cascading style). If present, the script uses it; 
          otherwise, it safely falls back to the standard Naskh style.

  D. SOUTH ASIAN (Indic Scripts / Abugidas)
     * Filenames:
       - "NotoSansDevanagari-Regular.ttf" (Hindi, Marathi)
       - "NotoSansBengali-Regular.ttf"    (Bengali)
       - "NotoSansGujarati-Regular.ttf"   (Gujarati)
       - "NotoSansTamil-Regular.ttf"      (Tamil)
       - "NotoSansTelugu-Regular.ttf"     (Telugu)
       - "NotoSansGurmukhi-Regular.ttf"   (Western Punjabi)
     * Source:    https://github.com/notofonts/noto-fonts (Download the "Phase 3" zip)
     * Why:       
       1. Complex Layout: These scripts use engines where vowels "float" above, below, or wrap 
          around consonants.
       2. Rendering: Without these specific fonts, vowels will detach from their consonants and 
          render as dotted circles (◌) or meaningless boxes.

  E. SOUTHEAST ASIAN
     * Filenames: "NotoSansThai-Regular.ttf", "NotoSansJavanese-Regular.ttf"
     * Why:       Thai tone marks must stack vertically at precise heights. Javanese is a rare 
                  historical script often completely missing from standard Windows/Mac systems.

====================================================================================================
"""

import pandas as pd
import os
import sys
import time
import threading
import re
import platform
from math import ceil

# ThreadPoolExecutor is essential for performance. Translating row-by-row sequentially
# is extremely slow. This allows us to translate multiple languages simultaneously.
from concurrent.futures import ThreadPoolExecutor, as_completed

# Deep Translator serves as the Python wrapper for the Google Translate API.
# It handles the HTTP requests and parsing of the translation response.
from deep_translator import GoogleTranslator

# OpenPyXL is the library used to manipulate Excel files directly.
# We use it specifically for the "Post-Processing" phase to apply font formatting.
from openpyxl import load_workbook
from openpyxl.styles import Font

# ==============================================================================
# SECTION A: GLOBAL CONFIGURATION & CONSTANTS
# ==============================================================================

# 1. FILE SYSTEM PATHS
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE  = os.path.join(BASE_DIR, "english.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "glossary.xlsx")
FONTS_DIR   = os.path.join(BASE_DIR, "fonts")

# 2. API PERFORMANCE SETTINGS
CHUNK_SIZE  = 50      # We batch 50 words into a single HTTP request to minimize overhead.
MAX_WORKERS = 1       # We restrict threads to 1. Increasing this increases the risk of IP bans.
REQUEST_DELAY = 1.5   # We intentionally pause for 1.5 seconds between batches to be "polite" to the API.

# 3. CONSOLE UI SETTINGS
BAR_WIDTH = 40        # The visual width (in characters) of the progress bar in the terminal.

# ==============================================================================
# SECTION B: FONT MAPPING CONFIGURATION
# ==============================================================================
# This dictionary is the "Brain" of the font selection engine.
# It maps internal Language Names to search terms and Excel Font Family names.
#
# 'search_terms': Expanded keywords to find the file in ./fonts OR System Fonts (Windows).
# 'family': The exact string to write into the Excel file properties.

FONT_CONFIG = {
    # --- 1. GLOBAL DEFAULTS ---
    # Used for languages that use standard Latin characters (English, French, etc.)
    "Default": {
        "search_terms": ["notosansliving", "notosans-regular", "arial", "helvetica", "calibri", "segoeui"],
        "family": "Noto Sans" # Falls back to system default if not installed, but NotoSansLiving is preferred
    },
    
    # --- 2. WESTERN / EUROPEAN ---
    "French": { "search_terms": ["notosans-regular", "arial"], "family": "Noto Sans" },
    "Spanish": { "search_terms": ["notosans-regular", "arial"], "family": "Noto Sans" },
    "German": { "search_terms": ["notosans-regular", "arial"], "family": "Noto Sans" },
    "Italian": { "search_terms": ["notosans-regular", "arial"], "family": "Noto Sans" },
    "Portuguese": { "search_terms": ["notosans-regular", "arial"], "family": "Noto Sans" },
    "Turkish": { "search_terms": ["notosans-regular", "arial"], "family": "Noto Sans" },
    "Indonesian": { "search_terms": ["notosansliving", "notosans-regular", "arial"], "family": "Noto Sans" },
    
    # --- 3. SPECIAL LATIN & CYRILLIC ---
    "Russian": { 
        "search_terms": ["notosanscyrillic", "notosans-regular", "arial", "segoeui"], 
        "family": "Noto Sans" 
    },
    "Vietnamese": { 
        "search_terms": ["notosansvietnamese", "notosans-regular", "arial", "segoeui"], 
        "family": "Noto Sans" 
    },
    "Hausa": { "search_terms": ["notosans-regular", "arial"], "family": "Noto Sans" },
    "Swahili": { "search_terms": ["notosans-regular", "arial"], "family": "Noto Sans" },

    # --- 4. EAST ASIAN (CJK) ---
    # These require the massive .ttc collections.
    "Mandarin_Chinese": {
        "search_terms": ["notosanssc", "notosanscjksc", "simhei", "simkai", "arialuni", "dengxian", "notoserifcjk"],
        "family": "Noto Sans SC"
    },
    "Wu_Chinese": { 
        "search_terms": ["notosanssc", "notosanscjksc", "simhei", "arialuni"],
        "family": "Noto Sans SC"
    },
    "Yue_Chinese": { 
        "search_terms": ["notosanstc", "notosanscjktc", "microsoftjhenghei", "msjh", "mingliu", "pmingliu", "simhei"],
        "family": "Noto Sans TC"
    },
    "Japanese": {
        "search_terms": ["notosansjp", "notosanscjkjp", "msgothic", "meiryo", "arialuni"],
        "family": "Noto Sans JP"
    },
    "Korean": {
        "search_terms": ["notosanskr", "notosanscjkkr", "malgun", "gulim", "arialuni"],
        "family": "Noto Sans KR"
    },

    # --- 5. MIDDLE EASTERN (RTL) ---
    "Arabic": { "search_terms": ["notosansarabic", "arial", "tahoma", "segoeui"], "family": "Noto Sans Arabic" },
    "Egyptian_Arabic": { "search_terms": ["notosansarabic", "arial"], "family": "Noto Sans Arabic" },
    "Iranian_Persian": { "search_terms": ["notosansarabic", "arial"], "family": "Noto Sans Arabic" },
    "Urdu": { 
        "search_terms": ["notonastaliqurdu", "notosansarabic", "arial", "tahoma"], 
        "family": "Noto Nastaliq Urdu" # Prefer Nastaliq, but Noto Sans Arabic is a valid fallback in Excel
    },
    
    # --- 6. SOUTH ASIAN (Indic Scripts) ---
    "Hindi": { "search_terms": ["notosansdevanagari", "mangal", "nirmala", "aparajita"], "family": "Noto Sans Devanagari" },
    "Marathi": { "search_terms": ["notosansdevanagari", "mangal"], "family": "Noto Sans Devanagari" },
    "Bengali": { "search_terms": ["notosansbengali", "vrinda"], "family": "Noto Sans Bengali" },
    "Telugu": { "search_terms": ["notosanstelugu", "gautami"], "family": "Noto Sans Telugu" },
    "Tamil": { "search_terms": ["notosanstamil", "latha"], "family": "Noto Sans Tamil" },
    "Gujarati": { "search_terms": ["notosansgujarati", "shruti"], "family": "Noto Sans Gujarati" },
    
    # Western Punjabi logic: Google often returns Gurmukhi. We search for Gurmukhi fonts too.
    "Western_Punjabi": { "search_terms": ["notosansarabic", "notosansgurmukhi", "raavi"], "family": "Noto Sans Gurmukhi" },

    # --- 7. SOUTHEAST ASIAN ---
    "Thai": { "search_terms": ["notosansthai", "leelawadee", "tahoma"], "family": "Noto Sans Thai" },
    "Javanese": { "search_terms": ["notosansjavanese", "notosans-regular", "javatext"], "family": "Noto Sans Javanese" },
}

# ==============================================================================
# SECTION C: ROBUST FONT ENGINE (Synchronized with script_glossary.py)
# ==============================================================================
# This section handles the physical discovery of fonts.
# It makes the script "portable" - it searches the folder rather than hardcoding paths.

FONT_PATH_CACHE = {} 

def normalize_name(name):
    """
    Normalizes a filename to make fuzzy matching easier.
    Input: "01_NotoSans-Regular.ttf" -> Output: "notosans"
    """
    name = name.lower()
    name = os.path.splitext(name)[0]
    name = re.sub(r'[\s\-_]', '', name)
    name = name.replace("regular", "")
    # Clean variable font junk if present
    name = name.replace("vf", "").replace("variablefont", "")
    name = re.sub(r'wght\d+', '', name)
    return name

def scan_directory(base_path, is_system=False):
    """
    Recursively scans a folder for .ttf, .otf, and .ttc files.
    Populates the global FONT_PATH_CACHE.
    """
    if not os.path.exists(base_path):
        return 0
    
    count = 0
    if is_system:
        try:
            files = os.listdir(base_path)
            walker = [(base_path, [], files)]
        except:
            return 0
    else:
        walker = os.walk(base_path)

    for root, dirs, files in walker:
        for f in files:
            f_lower = f.lower()
            if not f_lower.endswith(('.ttf', '.otf', '.ttc')):
                continue
            # Block Variable Fonts (Optional here, but good for consistency with PDF gen)
            if "-vf" in f_lower or "variable" in f_lower or "wght" in f_lower:
                continue

            full_path = os.path.join(root, f)
            norm_key = normalize_name(f)
            
            if norm_key not in FONT_PATH_CACHE:
                FONT_PATH_CACHE[norm_key] = []
            
            # Prioritize TTC files
            if f_lower.endswith('.ttc'):
                FONT_PATH_CACHE[norm_key].insert(0, full_path)
            else:
                FONT_PATH_CACHE[norm_key].append(full_path)
            count += 1
    return count

def scan_all_fonts():
    """
    Orchestrates the font scanning process (Local Folder + System Folder).
    """
    print(f"--- [System] Scanning Fonts... ---")
    local_count = scan_directory(FONTS_DIR)
    print(f"  > Local './fonts': Found {local_count} SAFE fonts.")
    
    sys_count = 0
    if platform.system() == "Windows":
        sys_font_dir = "C:\\Windows\\Fonts"
        print(f"  > Scanning Windows System Fonts: {sys_font_dir}...")
        try:
            sys_count = scan_directory(sys_font_dir, is_system=True)
            print(f"  > Windows System: Found {sys_count} SAFE fonts.")
        except Exception as e:
            print(f"  > [Warning] Could not scan Windows fonts: {e}")

    print(f"--- Scan complete. Total Index: {local_count + sys_count} fonts. ---\n")

def get_excel_font_family(language_name):
    """
    Determines the best Excel Font Family to apply to a specific column.
    
    The Logic Flow:
    1. Check FONT_CONFIG for the language's preferred settings.
    2. Check if the user physically possesses the required font file (via FONT_PATH_CACHE).
    3. If they have the file (Local or System), return the specific family name (e.g. "Noto Sans Thai").
    4. If they DO NOT have the file, return "Calibri" to avoid Excel errors.
    """
    # 1. Determine Config
    if language_name in FONT_CONFIG:
        config = FONT_CONFIG[language_name]
    else:
        # Fallback search
        config = FONT_CONFIG.get("Default")
        lang_search = language_name.replace(" ", "_")
        for key, val in FONT_CONFIG.items():
            if key in lang_search:
                config = val
                break
    
    # 2. Validation: Verify file existence via Cache
    search_terms = config.get('search_terms', [])
    
    for term in search_terms:
        term_norm = normalize_name(term)
        
        # Direct Match
        if term_norm in FONT_PATH_CACHE:
            return config['family']
            
        # Fuzzy Match against Cache Keys
        for cached_key in FONT_PATH_CACHE.keys():
            if term_norm in cached_key:
                # Success: We found the file (either locally or in Windows Fonts)
                return config['family']

    # 3. Fallback: If font file is missing, use safe system default.
    return "Calibri"

# ==============================================================================
# SECTION D: CORE TRANSLATION LOGIC
# ==============================================================================

# Thread-safe locks and counters for the console progress bar
progress_lock = threading.Lock()
total_chunks_global = 0
processed_chunks_global = 0

def update_progress_bar():
    """
    Updates the visual progress bar in the terminal.
    Uses '\r' (Carriage Return) to overwrite the current line, creating an animation.
    """
    global processed_chunks_global, total_chunks_global
    with progress_lock:
        processed_chunks_global += 1
        curr, tot = processed_chunks_global, total_chunks_global

    if tot == 0: pct = 0
    else: pct = (curr / tot) * 100
    
    fill = int(BAR_WIDTH * curr // tot) if tot > 0 else 0
    bar = '=' * fill + '-' * (BAR_WIDTH - fill)
    sys.stdout.write(f'\rTranslation Progress: [{bar}] {pct:.1f}%')
    sys.stdout.flush()

def batch_translate_text(text_list, target_code):
    """
    Orchestrates the API calls to Google Translate.
    """
    translator = GoogleTranslator(source='auto', target=target_code)
    results = []
    
    for i in range(0, len(text_list), CHUNK_SIZE):
        batch = text_list[i : i + CHUNK_SIZE]
        
        # Pre-process: Convert NaNs (empty cells) to empty strings so API doesn't choke
        clean_batch = [str(t) if pd.notna(t) and str(t).strip() else "" for t in batch]
        
        # Optimization: Don't call API if the whole batch is empty rows
        if all(x == "" for x in clean_batch):
            results.extend([""] * len(batch))
            update_progress_bar()
            continue
            
        try:
            # The main API call
            res = translator.translate_batch(clean_batch)
            results.extend(res)
        except Exception as e:
            # Fallback Strategy: If a batch fails (e.g. one malformed character),
            # switch to slow-mode and translate one by one to save the valid data.
            temp = []
            for item in clean_batch:
                try: 
                    if item == "": temp.append("")
                    else: temp.append(translator.translate(item))
                except: 
                    temp.append("[Translation Error]")
            results.extend(temp)
        
        update_progress_bar()
        time.sleep(REQUEST_DELAY) # Mandatory pause for API safety
        
    return results

def worker_process_language(df, lang_id, lang_def):
    """
    The Worker Thread Function.
    This runs in parallel for each selected language.
    """
    code = lang_def['code']
    name = lang_def['name']
    
    # 1. Normalize Column Header
    # 'script_glossary.py' expects underscores, no spaces, no parens.
    clean_name = name.replace(" ", "_").replace("(", "").replace(")", "")
    
    # 2. Translate 'Word' Column
    translated_words = batch_translate_text(df['English_word'].tolist(), code)
    
    # 3. Translate 'Description' Column
    # Check if 'English_descr' exists (it should, due to main() checks), otherwise default to empty.
    # This guarantees the description column is always created.
    if 'English_descr' in df.columns:
        translated_desc = batch_translate_text(df['English_descr'].tolist(), code)
    else:
        translated_desc = [""] * len(df)
    
    return {
        'w_col': f"{clean_name}_word",  # The column header for the word
        'w_data': translated_words,     # The list of translated strings
        'd_col': f"{clean_name}_descr", # The column header for the description
        'd_data': translated_desc       # The list of translated descriptions
    }

# ==============================================================================
# SECTION E: MAIN EXECUTION FLOW
# ==============================================================================

def main():
    global total_chunks_global
    
    print("\n========================================================")
    print("   HIGH-PERFORMANCE MULTILINGUAL EXCEL TRANSLATOR")
    print("========================================================")

    # 1. INITIALIZATION: Scan for fonts using Robust Engine
    scan_all_fonts()

    # 2. LANGUAGE DATABASE
    # Maps internal IDs to Google Translate (ISO 639-1) codes.
    langs = {
        1:  {"name": "English", "code": "en"},
        2:  {"name": "Mandarin Chinese", "code": "zh-CN"},
        3:  {"name": "Hindi", "code": "hi"},
        4:  {"name": "Spanish", "code": "es"},
        5:  {"name": "Portuguese", "code": "pt"}, 
        6:  {"name": "Standard Arabic", "code": "ar"}, 
        7:  {"name": "Bengali", "code": "bn"},
        8:  {"name": "French", "code": "fr"},
        9:  {"name": "Russian", "code": "ru"},
        10: {"name": "Urdu", "code": "ur"},
        11: {"name": "Indonesian", "code": "id"},
        12: {"name": "German", "code": "de"},
        13: {"name": "Japanese", "code": "ja"},
        14: {"name": "Marathi", "code": "mr"},
        15: {"name": "Telugu", "code": "te"},
        16: {"name": "Turkish", "code": "tr"},
        17: {"name": "Tamil", "code": "ta"},
        18: {"name": "Yue Chinese", "code": "zh-TW"}, # Traditional Chinese (Cantonese)
        19: {"name": "Wu Chinese", "code": "zh-CN"},  # Shanghainese (Uses Simplified)
        20: {"name": "Korean", "code": "ko"},
        21: {"name": "Vietnamese", "code": "vi"},
        22: {"name": "Hausa", "code": "ha"},
        23: {"name": "Iranian Persian", "code": "fa"},
        24: {"name": "Egyptian Arabic", "code": "ar"},
        25: {"name": "Swahili", "code": "sw"},
        26: {"name": "Javanese", "code": "jw"},
        27: {"name": "Italian", "code": "it"},
        28: {"name": "Western Punjabi", "code": "pa"}, 
        29: {"name": "Gujarati", "code": "gu"},
        30: {"name": "Thai", "code": "th"}
    }

    # 3. USER INTERFACE
    print(f"\n{'ID':<5} {'Language'}")
    print("-" * 30)
    for k in sorted(langs.keys()):
        print(f"{k:<5} {langs[k]['name']}")

    u_in = input("\nEnter IDs to translate (e.g. 2, 6, 13) or 'all': ")
    sel_ids = []
    
    if u_in.strip().lower() == 'all':
        sel_ids = list(langs.keys())
    else:
        try:
            parts = [int(x.strip()) for x in u_in.split(',')]
            for p in parts:
                if p in langs: sel_ids.append(p)
        except:
            print("Invalid input. Please enter numbers separated by commas.")
            return

    if not sel_ids: 
        print("No languages selected. Exiting.")
        return

    # 4. LOAD SOURCE DATA
    if not os.path.exists(INPUT_FILE):
        print(f"\n[ERROR] {INPUT_FILE} not found.")
        print("Please create an Excel file with 'English_word' and 'English_descr' columns.")
        return

    print(f"\nReading {INPUT_FILE}...")
    df = pd.read_excel(INPUT_FILE).fillna('')
    
    # ----------------------------------------------------------------------
    # FIX START: Ensure 'English_descr' exists to guarantee 2-column output
    # ----------------------------------------------------------------------
    if 'English_descr' not in df.columns:
        print("[INFO] 'English_descr' column missing in source. Creating empty column to ensure structure.")
        df['English_descr'] = ""
    # ----------------------------------------------------------------------

    target_ids = [i for i in sel_ids if i != 1]
    
    # 5. EXECUTE TRANSLATIONS
    if target_ids:
        row_count = len(df)
        total_chunks_global = ceil(row_count / CHUNK_SIZE) * 2 * len(target_ids)
        
        print(f"Translating {row_count} rows into {len(target_ids)} languages...")
        
        results = []
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = []
            for lid in target_ids:
                futures.append(executor.submit(worker_process_language, df, lid, langs[lid]))
            
            for future in as_completed(futures):
                try: results.append(future.result())
                except Exception as e: print(f"\nThread Error: {e}")

        sys.stdout.write(f"\rTranslation Progress: [{'=' * BAR_WIDTH}] 100.0%\n")
        
        print("Merging translated data columns...")
        for res in results:
            df[res['w_col']] = res['w_data']
            df[res['d_col']] = res['d_data']

    # 6. SAVE RAW DATA
    print(f"Saving raw data to '{OUTPUT_FILE}'...")
    df.to_excel(OUTPUT_FILE, index=False)

    # 7. TYPOGRAPHIC POST-PROCESSING
    print("Applying smart font formatting to columns...")
    try:
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        
        for col in ws.iter_cols():
            header = str(col[0].value)
            lang_name = None
            
            # Identify Language from Header
            if header.lower() == "category":
                lang_name = "Default"
            elif header.endswith("_word"):
                lang_name = header.replace("_word", "")
            elif header.endswith("_descr"):
                lang_name = header.replace("_descr", "")
            
            # If valid language column, find and apply the font
            if lang_name:
                font_family = get_excel_font_family(lang_name)
                
                # Apply only if we have a specific font family recommendation
                if font_family != "Calibri":
                    special_font = Font(name=font_family)
                    for cell in col[1:]: 
                        cell.font = special_font
                    print(f"   -> Applied '{font_family}' to column '{header}'")
        
        wb.save(OUTPUT_FILE)
        print("\nSUCCESS: 'glossary.xlsx' created with correct translations and fonts.")
        print("Next Step: Run 'script_glossary.py' to generate the PDF.")
        
    except Exception as e:
        print(f"[WARN] Could not apply fonts: {e}")

if __name__ == "__main__":
    main()